#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, colors
import time


class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excel_file = None
        # 设置字体的颜色
        self.font = Font(color=None)
        # 颜色对应的RGB值
        self.rgb_dict = {'red': colors.RED, 'green': colors.GREEN}

    def load_workbook(self, excel_path_and_name):
        """
        将Excel文件加载到内存，并获取其workbook对象
        :param excel_path_and_name: excel文件路径或名称
        :return: 返回workbook对象
        """
        try:
            self.workbook = openpyxl.load_workbook(excel_path_and_name)
        except Exception as e:
            raise e
        self.excel_file = excel_path_and_name
        return self.workbook

    def get_sheet_by_name(self, sheet_name):
        """
        根据sheet名获取该sheet对象
        :param sheet_name: excel表的sheet名称
        :return: 返回sheet对象
        """
        try:
            sheet = self.workbook[sheet_name]
            return sheet
        except Exception as e:
            raise e

    def get_sheet_by_index(self, sheet_index):
        """
        根据sheet的索引号获取该sheet对象，sheet_index从0开始
        :param sheet_index: sheet的索引号
        :return: 返回sheet对象
        """
        try:
            sheet_name = self.workbook.sheetnames[sheet_index]
        except Exception as e:
            raise e
        sheet = self.workbook[sheet_name]
        return sheet

    def get_rows_num(self, sheet):
        """
        获取sheet中有数据区域的结束行号
        :param sheet: sheet对象
        :return: 返回结束行号
        """
        num = sheet.max_row
        return num

    def get_cols_num(self, sheet):
        """
        获取sheet中有数据区域的结束列号
        :param sheet: sheet对象
        :return: 返回列号
        """
        return sheet.max_column

    def get_start_rows_num(self, sheet):
        """
        获取sheet中有数据区域的开始的行号,从1开始
        :param sheet: sheet对象
        :return: 返回行号
        """
        return sheet.min_row

    def get_start_cols_num(self, sheet):
        """
        获取sheet中有数据区域的开始的列号
        :param sheet: sheet对象
        :return: 返回列号
        """
        return sheet.min_column

    def get_row(self, sheet, row_no):
        """
        获取sheet中某一行，返回的是这一行所有的数据内容组成的元组
        下标从1开始，sheet.rows[1]表示第一行
        :param sheet: sheet对象
        :param row_no: 行号
        :return: 返回元组
        """
        try:
            rows = list(sheet.rows)
            rows = rows[row_no - 1]
            rows_list = [str(row.value).strip() for row in rows]
            return rows_list
        except Exception as e:
            raise e

    def get_column(self, sheet, col_no):
        """
        获取sheet中某一列，返回的是这一行所有的数据内容组成的元组
        小标从1开始，sheet.columns[1]标识第一列
        :param sheet: sheet对象
        :param col_no: 列号
        :return: 返回元组
        """
        try:
            columns = list(sheet.columns)
            columns = columns[col_no - 1]
            columns_list = [str(column.value).strip() for column in columns]
            return columns_list
        except Exception as e:
            raise e

    def get_cell_of_value(self, sheet, coordinate=None, row_no=None, col_no=None):
        """
        根据单元格所在的位置索引获取该单元格中的值，下标从1开始
        如果单元格里的没有值，则获取的值为None
        :param sheet: sheet对象
        :param coordinate:单元格编码坐标
        :param row_no: 行号
        :param col_no: 列号
        :return: 返回单元格的值
        """
        if coordinate:
            try:
                cell_value = sheet[coordinate].value
                # Excell表格中获取的值有可能为int类型，为了去空格，需要转化成字符串
                if isinstance(cell_value, int):
                    cell_value = str(cell_value)
                return cell_value.strip()
            except Exception as e:
                raise e
        elif coordinate is None and row_no and col_no:
            try:
                cell_value = sheet.cell(row=row_no, column=col_no).value
                return cell_value.strip()
            except Exception as e:
                raise e
        else:
            raise Exception('Insufficient Coordinates of cell!')

    def get_cell_of_object(self, sheet, coordinate=None, row_no=None, col_no=None):
        """
        获取某个单元格的对象，可以根据单元格所在位置的数字索引
        也可以直接根据Excel中单元格的编码及坐标获取
        如get_cell_of_object(sheet, coordinate='A1')或者 get_cell_of_object(sheet, row_no=1, col_no=2)
        :param sheet: sheet对象
        :param coordinate: 单元格的编码坐标
        :param row_no: 行号
        :param col_no: 列号
        :return: 返回单元格对象
        """
        if coordinate:
            try:
                return sheet.cell[coordinate]
            except Exception as e:
                raise e
        elif coordinate is None and row_no and col_no:
            try:
                return sheet.cell(row=row_no, column=col_no)
            except Exception as e:
                raise e
        else:
            raise Exception('Insufficient Coordinate of cell!')

    def write_cell(self, sheet, content, coordinate=None, row_no=None, col_no=None, style=None):
        """
        根据单元格在Excel中的编码坐标或者数字索引坐标向单元格中写入数据
        下标从1开始，参数style标识字体的颜色的名字，比如red，green
        :param sheet: sheet对象
        :param content: 单元格写入的内容
        :param coordinate: 单元格编码坐标
        :param row_no: 行号
        :param col_no: 列号
        :param style: 字体颜色名称
        """
        if coordinate:
            try:
                sheet.cell[coordinate].value = content
                if style:
                    sheet.cell[coordinate].font = Font(color=self.rgb_dict[style])
                self.workbook.save(self.excel_file)
            except Exception as e:
                raise e
        elif coordinate is None and row_no and col_no:
            try:
                sheet.cell(row=row_no, column=col_no).value = content
                if style:
                    sheet.cell(row=row_no, column=col_no).font = Font(color=self.rgb_dict[style])
                self.workbook.save(self.excel_file)
            except Exception as e:
                raise e
        else:
            raise Exception('Insufficient Coordinate of cell!')

    def write_cell_current_time(self, sheet, coordinate=None, row_no=None, col_no=None):
        """
        写入当前的时间，下标从1开始
        :param sheet: sheet对象
        :param coordinate: 单元格的编码坐标
        :param row_no: 行号
        :param col_no: 列号
        """
        # 显示为时间戳
        now = int(time.time())
        time_arr = time.localtime(now)
        current_time = time.strftime('%Y-%m-%d %H:%M:%S', time_arr)
        if coordinate:
            try:
                sheet.cell[coordinate].value = current_time
                self.workbook.save(self.excel_file)
            except Exception as e:
                raise e
        elif coordinate is None and row_no and col_no:
            try:
                sheet.cell(row=row_no, column=col_no).value = current_time
                self.workbook.save(self.excel_file)
            except Exception as e:
                raise e
        else:
            raise Exception('Insufficient Coordinates of cell!')


# if __name__ == '__main__':
#     pe = ParseExcel()
#     # 测试所用的Excel文件“163邮箱联系人.xlsx”
#     pe.load_workbook(r'D:\zdh\163邮箱联系人.xlsx')
#     print('通过名称获取sheet对象的名字：', pe.get_sheet_by_name('联系人').title)
#     print('通过索引获取sheet对象的名字：', pe.get_sheet_by_index(1).title)
#     i_sheet = pe.get_sheet_by_index(1)
#     print(type(i_sheet))
#     print('最大行号：', pe.get_rows_num(i_sheet))
#     print('最大列号：', pe.get_cols_num(i_sheet))
#     rows = pe.get_row(i_sheet, 1)
#     for i in rows:
#         print(i.value)
#     print('---------------------------------')
#     print('一行一列单元格的值：', pe.get_cell_of_value(i_sheet, coordinate='A1'))
#     # pe.write_cell(i_sheet, '刻舟求剑', row_no=10, col_no=10)
#     # pe.write_cell_current_time(i_sheet, row_no=11, col_no=10)
#     two_col = pe.get_column(i_sheet, 2)
#     for col in two_col:
#         print(col.value)
#     print('-----------------------------------')
#     i_sheet_0 = pe.get_sheet_by_index(0)
#     one_col = pe.get_column(i_sheet_0, 4)
#     for one in one_col:
#         print(one.value)

if __name__ == '__main__':
    pe = ParseExcel()
    pe.load_workbook('..//testdata//163邮箱发送邮件.xlsx')
    i_sheet = pe.get_sheet_by_index(1)
    row_num = pe.get_row(i_sheet, 2)
    print(row_num)
    print(pe.get_rows_num(i_sheet))
    print(pe.get_start_rows_num(i_sheet))
    print(pe.get_cols_num(i_sheet))
    print(pe.get_start_cols_num(i_sheet))
    # pe.write_cell_current_time(i_sheet, row_no=7, col_no=9)
    # pe.write_cell(i_sheet, 'pass', row_no=7, col_no=10, style='green')
